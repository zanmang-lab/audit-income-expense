"""업로드 파일만으로 결과가 나오는지 검증 (config/학기설정.json 미사용)."""
from __future__ import annotations

import shutil
import sys
import tempfile
from io import BytesIO
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.parse_overdue_rules_image import parse_overdue_rules_image
from src.pipeline import run_engine
from src.remarks import parse_overdue_sheet, parse_refund_sheet
from src.semester_config import semester_config_to_dict
from src.parse_ledger import load_ledger_workbook, parse_ledger_workbook
from web.masking import mask_refund_workbook
from web.semester_inputs import build_semester_from_overview

CONFIG_PATH = PROJECT_ROOT / "config" / "학기설정.json"
BACKUP_PATH = PROJECT_ROOT / "config" / "학기설정.json.bak_verify"

# config와 완전히 다른 업로드 데이터
CUSTOM_RULES_SVG = """<svg xmlns="http://www.w3.org/2000/svg">
  <text x="10" y="20">테스트물품A — 5,000원/일</text>
  <text x="10" y="40">그 외 물품 — 777원/일 (기본)</text>
</svg>"""

BUSINESS_OVERVIEW = """77. 검증용물품대여사업
88. 검증용사물함사업"""


def _build_ledger_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "검증장부"
    ws.append(
        ["날짜", "사업번호", "세부번호", "기재내용", "현금", "구분", "개요", "수입", "지출", "오류여부"]
    )
    ws.append(
        [
            "2026-05-01",
            77,
            1,
            "홍길동",
            "",
            "지출의환급",
            "검증용 물품 대여 연체료",
            10000,
            0,
            "",
        ]
    )
    ws.append(
        [
            "2026-05-02",
            88,
            1,
            "전자99김철수",
            "",
            "지출의환급",
            "검증용 사물함 수익금",
            4000,
            0,
            "",
        ]
    )
    ws.append(
        [
            "2026-05-03",
            88,
            2,
            "이순신",
            "",
            "사업진행비",
            "검증용 사물함 환불",
            0,
            4000,
            "",
        ]
    )
    buf = BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _build_overdue_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "연체료목록"
    ws.append(["반납일(입금일)", "대여물품", "성명", "연체료", "비고"])
    ws.append(["2026-05-01", "테스트물품A 1개", "홍길동", 10000, ""])
    buf = BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _build_refund_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "환불자목록"
    ws.append(["사물함 번호", "이름", "환불금액", "사유"])
    ws.append(["전자-99", "이순신", 4000, "검증환불"])
    buf = BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _hide_config() -> None:
    if CONFIG_PATH.exists():
        shutil.move(CONFIG_PATH, BACKUP_PATH)


def _restore_config() -> None:
    if BACKUP_PATH.exists():
        shutil.move(BACKUP_PATH, CONFIG_PATH)


def main() -> int:
    errors: list[str] = []
    _hide_config()
    try:
        rules = parse_overdue_rules_image(
            CUSTOM_RULES_SVG.encode("utf-8"), "rules.svg"
        )
        if rules.default_per_day != 777:
            errors.append(f"기본 단가: 기대 777, 실제 {rules.default_per_day}")
        kw = [k for c in rules.categories for k in c.keywords]
        if "테스트물품A" not in kw:
            errors.append(f"이미지 키워드 누락: {kw}")

        semester = build_semester_from_overview(BUSINESS_OVERVIEW, rules)
        if semester.name_for(77) != "검증용물품대여사업":
            errors.append("사업개요 사업명 미반영")
        if semester.role_for(77) != "rental":
            errors.append("대여 역할 추론 실패")
        if semester.overdue_rules.default_per_day != 777:
            errors.append("학기설정에 연체료 규칙이 섞임")

        ledger_wb = load_ledger_workbook(BytesIO(_build_ledger_xlsx()))
        try:
            ledger = parse_ledger_workbook(ledger_wb, semester=semester)
        finally:
            ledger_wb.close()

        overdue_wb = openpyxl.load_workbook(BytesIO(_build_overdue_xlsx()), data_only=True)
        try:
            overdue = parse_overdue_sheet(overdue_wb.active)
        finally:
            overdue_wb.close()

        refund_bytes, _ = mask_refund_workbook(_build_refund_xlsx())
        refund_wb = openpyxl.load_workbook(BytesIO(refund_bytes), data_only=True)
        try:
            refund = parse_refund_sheet(refund_wb.active)
        finally:
            refund_wb.close()

        if not refund or refund[0].name != "이순신":
            errors.append("환불자 목록 파싱 실패")

        result = run_engine(
            ledger_by_business=ledger,
            semester=semester,
            overdue_records=overdue,
            refund_records=refund,
        )

        rental = result.businesses.get(77)
        if not rental:
            errors.append("사업 77 결과 없음")
        else:
            remarks = " ".join(p.remark for p in rental.pages if p.remark)
            if "2일" not in remarks:
                errors.append(f"연체일수 2일 기대 (10000/5000), 비고: {remarks!r}")
            if "5,000" in remarks or "5000" in remarks:
                pass  # optional
            if "검증용물품대여사업" not in rental.business_name:
                errors.append(f"출력 사업명: {rental.business_name!r}")

        locker = result.businesses.get(88)
        if not locker:
            errors.append("사업 88 결과 없음")
        else:
            locker_remarks = " ".join(p.remark for p in locker.pages if p.remark)
            if "이순신" not in locker_remarks:
                errors.append(f"환불 비고(업로드 환불자목록) 미반영: {locker_remarks!r}")
            if "검증용사물함사업" not in locker.business_name:
                errors.append(f"사물함 사업명: {locker.business_name!r}")

        extracted = semester_config_to_dict(semester)["overdue_rules"]
        if extracted["default_per_day"] != 777:
            errors.append("추출 JSON에 config 규칙이 섞임")

        # config 파일 없을 때 웹 빌더가 서버 설정을 끌어오지 않는지
        try:
            from src.semester_config import default_semester_config

            default_semester_config()
            errors.append("config 없을 때 default_semester_config()가 성공함 (웹 경로 오염 가능)")
        except FileNotFoundError:
            pass

    finally:
        _restore_config()

    if errors:
        print("FAIL")
        for e in errors:
            print(" -", e)
        return 1

    print("PASS: 업로드(사업개요·규정이미지·장부·연체료·환불)만으로 결과 생성됨")
    print("  - 연체료 10,000원 / 이미지 규칙 5,000원/일 → 비고에 2일 연체")
    print("  - config/학기설정.json 없이도 동일하게 동작")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
