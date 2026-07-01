"""업로드용 예시 파일 생성 (장부·연체료·환불자 목록)."""
from __future__ import annotations

from pathlib import Path

import openpyxl

LEDGER_SHEET = "장부"

ALLOWED_EXAMPLES = frozenset(
    {
        "장부_예시.xlsx",
        "연체료목록_예시.xlsx",
        "환불자목록_예시.xlsx",
    }
)


def ensure_example_files(examples_dir: Path) -> None:
    examples_dir.mkdir(parents=True, exist_ok=True)
    _write_ledger_sample(examples_dir / "장부_예시.xlsx")
    _write_overdue_sample(examples_dir / "연체료목록_예시.xlsx")
    _write_refund_sample(examples_dir / "환불자목록_예시.xlsx")


def resolve_example_path(examples_dir: Path, filename: str) -> Path | None:
    if filename not in ALLOWED_EXAMPLES:
        return None
    path = examples_dir / filename
    if not path.exists():
        ensure_example_files(examples_dir)
    return path if path.exists() else None


def _write_ledger_sample(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = LEDGER_SHEET
    ws.append(
        [
            "날짜",
            "사업번호",
            "세부번호",
            "기재내용",
            "현금",
            "구분",
            "개요",
            "수입",
            "지출",
            "오류여부",
        ]
    )
    ws.append(
        [
            "2026-03-10",
            2,
            1,
            "전자49장문경",
            "",
            "지출의환급",
            "1학기 대면 사물함 배부 사업 수익금 입금",
            4000,
            0,
            "",
        ]
    )
    ws.append(
        [
            "2026-03-23",
            3,
            1,
            "김지민",
            "",
            "지출의환급",
            "1학기 물품 대여 사업",
            4000,
            0,
            "",
        ]
    )
    wb.save(path)
    wb.close()


def _write_overdue_sample(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "연체료목록"
    ws.append(["반납일(입금일)", "대여물품", "성명", "연체료", "비고"])
    ws.append(["2026-03-23", "보조배터리 1개", "김지민", 4000, ""])
    ws.append(["2026-03-31", "폴라로이드 필름", "하명진", 7000, "필름값"])
    ws.append(["2026-04-01", "돗자리(대) 1개", "서윤아", 9000, "조소영'으로 입금"])
    wb.save(path)
    wb.close()


def _write_refund_sample(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "환불자목록"
    ws.append(["사물함 번호", "이름", "환불금액", "사유"])
    ws.append(["전자-12", "이지원", 4000, "당첨등록포기"])
    ws.append(["일반-05", "박민수", 2000, "두번송금"])
    wb.save(path)
    wb.close()
