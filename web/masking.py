"""환불자 목록 개인정보 열 마스킹 (web 전처리 전용)."""
from __future__ import annotations

import re
from io import BytesIO
from typing import Any

import openpyxl
from openpyxl import Workbook

# 엔진(parse_refund_sheet)이 기대하는 열 위치: 0=사물함, 1=이름, 3=금액, 4=사유
ENGINE_COL_LOCKER = 0
ENGINE_COL_NAME = 1
ENGINE_COL_AMOUNT = 3
ENGINE_COL_REASON = 4

SENSITIVE_HEADER_KEYWORDS = (
    "전화",
    "연락",
    "휴대",
    "핸드폰",
    "계좌",
    "phone",
    "tel",
    "mobile",
    "account",
)

PHONE_RE = re.compile(r"^01[016789]-?\d{3,4}-?\d{4}$")
ACCOUNT_RE = re.compile(r"^\d{8,}$")


def _norm_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _header_is_sensitive(header: str) -> bool:
    h = header.lower()
    return any(kw in h for kw in SENSITIVE_HEADER_KEYWORDS)


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _column_pattern_sensitive(values: list[str], threshold: float = 0.5) -> bool:
    non_empty = [v for v in values if v]
    if len(non_empty) < 2:
        return False
    phone_hits = sum(1 for v in non_empty if PHONE_RE.match(v.replace(" ", "")))
    account_hits = sum(
        1
        for v in non_empty
        if ACCOUNT_RE.match(re.sub(r"[^\d]", "", v)) and len(re.sub(r"[^\d]", "", v)) >= 8
    )
    return (phone_hits / len(non_empty) >= threshold) or (
        account_hits / len(non_empty) >= threshold
    )


def _semantic_column_map(headers: list[str]) -> dict[str, int | None]:
    mapping: dict[str, int | None] = {
        "locker": None,
        "name": None,
        "amount": None,
        "reason": None,
    }
    for idx, header in enumerate(headers):
        h = _norm_header(header)
        if not h:
            continue
        if mapping["locker"] is None and "사물함" in h:
            mapping["locker"] = idx
        elif mapping["name"] is None and "이름" in h:
            mapping["name"] = idx
        elif mapping["amount"] is None and ("환불" in h or "금액" in h):
            mapping["amount"] = idx
        elif mapping["reason"] is None and "사유" in h:
            mapping["reason"] = idx
    return mapping


def detect_sensitive_columns(headers: list[str], data_rows: list[tuple]) -> list[str]:
    """드롭 대상 열의 헤더명 목록."""
    return detect_sensitive_columns_from_rows(headers, data_rows)


def mask_refund_workbook(source: bytes) -> tuple[bytes, list[str]]:
    """
    환불자 목록 바이트를 읽어 민감 열을 제거한 뒤,
    엔진이 기대하는 열 위치(0,1,3,4)로 재구성한 워크북 바이트를 반환한다.
    """
    wb_in = openpyxl.load_workbook(BytesIO(source), read_only=True, data_only=True)
    ws = wb_in.active
    rows = list(ws.iter_rows(values_only=True))
    wb_in.close()

    if not rows:
        out = Workbook()
        buf = BytesIO()
        out.save(buf)
        return buf.getvalue(), []

    headers = [_cell_text(h) for h in rows[0]]
    dropped_names = detect_sensitive_columns_from_rows(headers, rows[1:])
    semantic = _semantic_column_map(headers)

    wb_out = Workbook()
    out_ws = wb_out.active
    out_ws.append(["사물함 번호", "이름", None, "환불금액", "사유", None])

    for row in rows[1:]:
        if not any(row):
            continue
        locker = _value_at(row, semantic["locker"])
        name = _value_at(row, semantic["name"])
        amount = _value_at(row, semantic["amount"])
        reason = _value_at(row, semantic["reason"])
        if not name:
            continue
        normalized = [None] * 6
        normalized[ENGINE_COL_LOCKER] = locker
        normalized[ENGINE_COL_NAME] = name
        normalized[ENGINE_COL_AMOUNT] = amount
        normalized[ENGINE_COL_REASON] = reason
        out_ws.append(normalized)

    buf = BytesIO()
    wb_out.save(buf)
    return buf.getvalue(), dropped_names


def _value_at(row: tuple, index: int | None) -> Any:
    if index is None or index >= len(row):
        return None
    return row[index]


def detect_sensitive_columns_from_rows(
    headers: list[str], data_rows: list[tuple]
) -> list[str]:
    dropped: list[str] = []
    dropped_indices: set[int] = set()
    for idx, header in enumerate(headers):
        if header and _header_is_sensitive(header):
            dropped_indices.add(idx)
            dropped.append(header)
    for idx, header in enumerate(headers):
        if idx in dropped_indices:
            continue
        col_values = [_cell_text(r[idx]) if idx < len(r) else "" for r in data_rows]
        if _column_pattern_sensitive(col_values):
            label = header or f"열{idx + 1}"
            if label not in dropped:
                dropped.append(label)
    return dropped
