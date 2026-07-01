"""장부 파싱 및 페이지 분할 (SPEC 3번)."""
from __future__ import annotations

from dataclasses import dataclass, field
from typing import Any, BinaryIO, Iterator

import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .semester_config import SemesterConfig

MAX_ROWS_PER_PAGE = 13


@dataclass
class LedgerRow:
    date: Any
    business_no: int
    detail_no: Any
    item_text: str
    overview: str | None
    income: float
    expense: float
    has_error: bool
    classification: str = ""
    classification_needs_review: bool = False


@dataclass
class Page:
    date: Any
    income_expense: str
    classification: str
    rows: list[LedgerRow] = field(default_factory=list)

    @property
    def key(self) -> tuple[Any, str, str]:
        return (self.date, self.income_expense, self.classification)


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _is_valid_business_no(value: Any) -> bool:
    if value is None:
        return False
    try:
        int(value)
        return True
    except (TypeError, ValueError):
        return False


def iter_ledger_rows(ws: Worksheet) -> Iterator[LedgerRow]:
    """워크시트에서 유효한 장부 행을 순서대로 반환."""
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        if not _is_valid_business_no(row[1]):
            continue

        income = float(row[7] or 0)
        expense = float(row[8] or 0)
        if income <= 0 and expense <= 0:
            continue

        overview = _cell_str(row[6]) if row[6] is not None else None
        classification = _cell_str(row[5]).strip() if len(row) > 5 and row[5] is not None else ""
        needs_review = not classification

        yield LedgerRow(
            date=row[0],
            business_no=int(row[1]),
            detail_no=row[2],
            item_text=_cell_str(row[3]),
            overview=overview,
            income=income,
            expense=expense,
            has_error=row[9] is not None and str(row[9]).strip() != "",
            classification=classification,
            classification_needs_review=needs_review,
        )


def _count_ledger_rows(ws: Worksheet) -> int:
    return sum(1 for _ in iter_ledger_rows(ws))


def find_ledger_sheet(wb: Workbook) -> str:
    """업로드 장부에서 데이터가 있는 시트를 자동 선택."""
    best_name = ""
    best_count = 0
    for name in wb.sheetnames:
        try:
            count = _count_ledger_rows(wb[name])
        except Exception:
            count = 0
        if count > best_count:
            best_count = count
            best_name = name

    if best_count > 0:
        return best_name

    raise ValueError(
        "장부 데이터를 찾지 못했습니다. "
        "날짜·사업번호·수입/지출 열이 있는 시트가 포함된 xlsx인지 확인해 주세요."
    )


def resolve_ledger_sheet(wb: Workbook, preferred: str | None = None) -> str:
    """선호 시트명이 있으면 먼저 시도하고, 없거나 비어 있으면 자동 탐색."""
    if preferred and preferred in wb.sheetnames and _count_ledger_rows(wb[preferred]) > 0:
        return preferred
    return find_ledger_sheet(wb)


def parse_ledger_workbook(
    wb: Workbook,
    sheet_name: str | None = None,
    semester: SemesterConfig | None = None,
) -> dict[int, list[LedgerRow]]:
    """열린 워크북에서 사업번호별 장부 행 목록을 반환."""
    preferred = sheet_name
    if semester is not None and preferred is None:
        preferred = semester.ledger_sheet or None
    resolved = resolve_ledger_sheet(wb, preferred)
    ws = wb[resolved]
    grouped: dict[int, list[LedgerRow]] = {}
    for ledger_row in iter_ledger_rows(ws):
        grouped.setdefault(ledger_row.business_no, []).append(ledger_row)
    return grouped


def split_pages(rows: list[LedgerRow]) -> list[Page]:
    """사업번호 내 행을 페이지 키 기준으로 분할. 키가 같아도 13줄이면 새 페이지."""
    pages: list[Page] = []
    current: Page | None = None

    for row in rows:
        income_expense = "수입" if row.income > 0 else "지출"
        key = (row.date, income_expense, row.classification)

        if (
            current is None
            or current.key != key
            or len(current.rows) >= MAX_ROWS_PER_PAGE
        ):
            current = Page(
                date=row.date,
                income_expense=income_expense,
                classification=row.classification,
            )
            pages.append(current)
        current.rows.append(row)

    return pages


def load_ledger_workbook(source: str | BinaryIO, read_only: bool = True) -> Workbook:
    """파일 경로 또는 바이너리 스트림에서 장부 워크북을 연다."""
    return openpyxl.load_workbook(source, read_only=read_only, data_only=True)
